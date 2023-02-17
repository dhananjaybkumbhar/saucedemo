import time

from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By

import pageObjects
from pageObjects.Login_page import driver
from utilities import excel
import openpyxl
from pageObjects import Login_page
import pytest as pytest
import xlrd
from utilities import Baseclass
from utilities.Baseclass import selenium

selenium()

def geturl():
    driver.get('https://www.saucedemo.com')



filename='C://Users/DELL/PycharmProjects/saucedemo/pageObjects/Testdata.xlsx'
global workbook
workbook=openpyxl.load_workbook(filename)
global sheet
sheet=workbook['Login']
global username
username=(sheet.cell(row=2,column=1).value)
global password
password=(sheet.cell(row=2,column=2).value)





@ pytest.mark.login

def test_S_Login_01():
    try:
        driver.find_element(By.ID, 'user-name').click()
        assert driver.find_element(By.ID, 'user_name').is_selected()
        driver.find_element(By.ID, 'password').click()
        assert driver.find_element(By.ID, 'password').is_selected()


    except NoSuchElementException:
        print('Exception Handled')
    time.sleep(2)

def test_S_Login_02():
    try:
        btn = driver.find_element(By.ID, 'login-button').click()
        time.sleep(5)
        msg = driver.find_element(By.CSS_SELECTOR,'#login_button_container > div > form > div.error-message-container.error').text
        assert msg=='Epic sadface: Username is required'

    except NoSuchElementException:
        print('Exception Handled')

    time.sleep(2)

#
def test_S_Login_03():
    try:
        driver.find_element(By.NAME, 'user-name').send_keys(username)
        driver.find_element(By.ID, 'password').send_keys(password)
        login_btn = driver.find_element(By.ID, 'login-button')
        login_btn.click()

        expected_url = 'https://www.saucedemo.com/inventory.html'
        actual_url = (driver.current_url)
        assert actual_url == expected_url

    except NoSuchElementException:
        print('Exception Handled')
    time.sleep(2)
#
def test_S_Login_04():
    geturl()
    username = (sheet.cell(row=3, column=1).value)
    password = (sheet.cell(row=3, column=2).value)

    try:
        driver.find_element(By.NAME, 'user-name').send_keys(username)
        driver.find_element(By.ID, 'password').send_keys(password)
        login_btn = driver.find_element(By.ID, 'login-button')
        login_btn.click()
        msg = driver.find_element(By.CSS_SELECTOR,'#login_button_container > div > form > div.error-message-container.error').text
        assert msg == 'Epic sadface: Username and password do not match any user in this service'

    except NoSuchElementException:
        print('Exception Handled')
    time.sleep(2)

def test_S_Login_05():
    geturl()
    username = (sheet.cell(row=4, column=1).value)
    password = (sheet.cell(row=4, column=2).value)

    try:
        driver.find_element(By.NAME, 'user-name').send_keys(username)
        driver.find_element(By.ID, 'password').send_keys(password)
        login_btn = driver.find_element(By.ID, 'login-button')
        login_btn.click()
        msg = driver.find_element(By.CSS_SELECTOR,'#login_button_container > div > form > div.error-message-container.error').text
        assert msg == 'Epic sadface: Username and password do not match any user in this service'

    except NoSuchElementException:
        print('Exception Handled')
    time.sleep(2)

def test_S_Login_06():
    geturl()
    username = (sheet.cell(row=5, column=1).value)
    password = (sheet.cell(row=5, column=2).value)

    try:
        driver.find_element(By.ID,'user-name').send_keys(username)
        driver.find_element(By.ID, 'password').send_keys(password)
        login_btn = driver.find_element(By.ID, 'login-button')
        login_btn.click()
        msg = driver.find_element(By.CSS_SELECTOR,'#login_button_container > div > form > div.error-message-container.error').text
        assert msg == 'Epic sadface: Username and password do not match any user in this service'

    except NoSuchElementException:
        print('Exception Handled')

    time.sleep(2)

def test_S_Login_07():
    try:
        driver.find_element(By.ID, 'user-name').send_keys('')
        driver.find_element(By.ID,'password').send_keys('')
        driver.find_element(By.ID,'login-button').click()
        msg = driver.find_element(By.CSS_SELECTOR,'#login_button_container > div > form > div.error-message-container.error').text
        assert msg == 'Epic sadface: Username and password do not match any user in this service'
        driver.get_screenshot_as_file('"C:/Users/DELL/Desktop/New/test_S_Login_07.png')

    except NoSuchElementException:
        print('Exception Handled')
